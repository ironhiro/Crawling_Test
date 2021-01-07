from bs4 import BeautifulSoup as bs
from PIL import Image
from io import BytesIO
import requests
import time
import openpyxl
from datetime import date
import os


def save_file(wb):
    today = date.today()
    if not os.path.isdir(today.isoformat()):
        os.mkdir(today.isoformat())
    print(wb.sheetnames)
    
    wb.remove(wb["Sheet"])
    wb.save("{}/sales_new.xlsx".format(today.isoformat()))
     


def do_coupang(url, wb):
    
    r = requests.get(url=url,headers=headers)
    if r.status_code == 200:
        soup = bs(r.text, 'html.parser')
        goods_thumbnail_list = soup.select('a > dl > dt > img')
        goods_discription_name = soup.select('a > dl > dd > div > div.name')
        goods_discription_value = soup.select('a > dl > dd > div > div.price-area > div > div.price > em > strong')

        print(len(goods_thumbnail_list))

        
        
        if url.find('latestAsc') != -1:
            ws = wb.create_sheet("coupang_최신순")
        else:
            ws = wb.create_sheet("coupang_판매량순")


        ws.append(['썸네일', '상품명', '가격'])
        cnt = 0
        for i,j,k in zip(goods_thumbnail_list, goods_discription_name, goods_discription_value):
            src = i.get('data-img-src')

            if src != None:
                print(src)
                if src.find('http')==-1:
                    src = 'https:' + src
                response = requests.get(src)
                src = openpyxl.drawing.image.Image(BytesIO(response.content))
                src.height = 100
                src.width = 100
                
                ws.add_image(src,str(chr(65)) + str(cnt+1+1))
                
                ws[str(chr(65+1)) + str(cnt+1+1)] = j.text
                
                ws[str(chr(65+2)) + str(cnt+1+1)] = k.text
                ws.row_dimensions[cnt+1+1].height=75
            else:
                src = '없음'
                ws[str(chr(65)) + str(cnt+1+1)] = src
                ws[str(chr(65+1)) + str(cnt+1+1)] = j.text
                ws[str(chr(65+2)) + str(cnt+1+1)] = k.text
            cnt+=1
        ws.column_dimensions['B'].width=125
        ws.column_dimensions['A'].width=12.5
        

def do_naver(url, wb):
    r = requests.get(url=url,headers=headers)
    if r.status_code == 200:
        soup = bs(r.text, 'html.parser')
        goods_thumbnail_list = soup.select('ul > li > div > img')
        goods_discription_name = soup.select('ul > li > div > em')
        goods_discription_value = soup.select('ul > li > div > strong')

        

        
        if url.find('UV') != -1:
            ws = wb.create_sheet("naver_많이본상품")
        else:
            ws = wb.create_sheet("naver_많이구매한상품")

        ws.append(['썸네일', '상품명', '가격'])
        cnt = 0
        for i,j,k in zip(goods_thumbnail_list, goods_discription_name, goods_discription_value):
            src = i.get('data-src')


            if src != None:
                print(src)    
                if src.find('http')==-1:
                    src = 'https:' + src
                response = requests.get(src)
                src = openpyxl.drawing.image.Image(BytesIO(response.content))
                src.height = 100
                src.width = 100
                ws.add_image(src,str(chr(65)) + str(cnt+1+1))
                ws[str(chr(65+1)) + str(cnt+1+1)] = j.text
                ws[str(chr(65+2)) + str(cnt+1+1)] = k.text.split('원')[0]
                ws.row_dimensions[cnt+1+1].height=75
            else:
                src = '없음'
                ws[str(chr(65)) + str(cnt+1+1)] = src
                ws[str(chr(65+1)) + str(cnt+1+1)] = j.text
                ws[str(chr(65+2)) + str(cnt+1+1)] = k.text.split('원')[0]
            cnt+=1
        ws.column_dimensions['B'].width=125
        ws.column_dimensions['A'].width=12.5

        

def do_11st(url, wb):
    r = requests.get(url=url,headers=headers)
    if r.status_code == 200:
        soup = bs(r.text, 'html.parser')
        goods_thumbnail_list = soup.select('div > ul > li > div > a > div.img_plot > img')
        goods_discription_name = soup.select('div > ul > li > div > a > div.pname > p')
        goods_discription_value = soup.select('div > ul > li > div > a > div.pname > div > span.price_detail > strong.sale_price')

        print(len(goods_thumbnail_list))

        
        ws = wb.create_sheet("11번가")
        ws.append(['썸네일', '상품명', '가격'])
        cnt = 0
        for i,j,k in zip(goods_thumbnail_list, goods_discription_name, goods_discription_value):
            src = i.get('src')

            if src != None:
                print(src)
                if src.find('http')==-1:
                    src = 'https:' + src
                response = requests.get(src)
                src = openpyxl.drawing.image.Image(BytesIO(response.content))
                src.height = 100
                src.width = 100
                ws.add_image(src,str(chr(65)) + str(cnt+1+1))
                ws[str(chr(65+1)) + str(cnt+1+1)] = j.text
                ws[str(chr(65+2)) + str(cnt+1+1)] = k.text
                ws.row_dimensions[cnt+1+1].height=75
                
            else:
                src = '없음'
                ws[str(chr(65)) + str(cnt+1+1)] = src
                ws[str(chr(65+1)) + str(cnt+1+1)] = j.text
                ws[str(chr(65+2)) + str(cnt+1+1)] = k.text
            cnt+=1
        ws.column_dimensions['B'].width=125
        ws.column_dimensions['A'].width=12.5
       

def do_gmarket(url, wb):
    r = requests.get(url=url,headers=headers)
    if r.status_code == 200:
        soup = bs(r.text, 'html.parser')
        goods_thumbnail_list = soup.select('div.best-list > ul > li > div.thumb > a > img.lazy')[:100]
        goods_discription_name = soup.select('div.best-list > ul > li > a.itemname')[:105]
        goods_discription_value = soup.select('div.best-list > ul > li > div.item_price')[:100]
        


        goods_discription_name = goods_discription_name[5:]
        
        
        ws = wb.create_sheet("지마켓")
        ws.append(['썸네일', '상품명', '가격'])
        cnt = 0
        
        for i,j,k in zip(goods_thumbnail_list, goods_discription_name, goods_discription_value):
            src = i.get('data-original')            
            
            if src != None:
                print(src)
                if src.find('http')==-1:
                    src = 'https:' + src
                response = requests.get(src)
                src = openpyxl.drawing.image.Image(BytesIO(response.content))
                src.height = 100
                src.width = 100
                ws.add_image(src,str(chr(65)) + str(cnt+1+1))
                ws[str(chr(65+1)) + str(cnt+1+1)] = j.text
                ws[str(chr(65+2)) + str(cnt+1+1)] = k.text.split('원')[0].replace('\n','')
                ws.row_dimensions[cnt+1+1].height=75
            else:
                src = '없음'
                ws[str(chr(65)) + str(cnt+1+1)] = src
                ws[str(chr(65+1)) + str(cnt+1+1)] = j.text
                ws[str(chr(65+2)) + str(cnt+1+1)] = k.text.split('원')[0].replace('\n','')
            cnt+=1
        ws.column_dimensions['B'].width=125
        ws.column_dimensions['A'].width=12.5
        
    
        





if __name__ == "__main__":

    with open('crawl_list.txt', encoding='utf-8') as f:
        s = f.readlines()
    wb = openpyxl.Workbook()
    
    
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Whale/2.8.108.15 Safari/537.36'}
    for i in s:
        name, url = i.split()
        if name == '네이버':
            do_naver(url, wb)
        elif name == '쿠팡':
            do_coupang(url, wb)
        elif name == '지마켓':
            do_gmarket(url, wb)
        else:
            do_11st(url, wb)


        
                
        time.sleep(2)
    save_file(wb)
    
        
    
        

   
    